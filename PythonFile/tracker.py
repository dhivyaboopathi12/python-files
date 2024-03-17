import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Variables for the Excel file and email information
lst = openpyxl.load_workbook('D:\\details.xlsx')

sheet = lst['sheet1']
row_size = sheet.max_row
col_size = sheet.max_column

l1 =[]     #list of stud
l2 = []    #append roll with attendaance
l3 = []    #rol id with lack of attendance

# mails ids
staff_mails = ['dhivyacds12@gmail.com','dhivyacds12@gmail.com']

#message
m1 = "Warnings!! you can take only one more day leave for CI class"
m2 = "Warnings!! you can take only one more day leave for pthon"
m3 = "Warnings!! you can take only one more day leave for data mining"

def savefile():
    lst.save(r'D:\\details.xlsx')
    print("Saved!!")

def check(days,row_num,sub):
    global staff_mails
    global l2
    global l3
    
    for std in range(0,len(row_num)):
        if days[std] is 2:
            if sub is 1:
                l1.append(sheet.cell(row_num[std],column=2).value)   
                mailstu(l1.m1)
            elif b is 2:
                l1.append(sheet.cell(row_num[std],column=2).value)   
                mailstu(l1.m12)
            else:
                l1.append(sheet.cell(row_num[std],column=2).value)   
                mailstu(l1.m3)   
        elif days > 2:
            if b is 1:
                l2 = l2+str(sheet.cell(row=row_num[std],column=1).value)
                l3.append(sheet)        